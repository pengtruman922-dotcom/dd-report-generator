"""Upload endpoints: Excel file, PDF/MD attachments, and manual input."""

from __future__ import annotations

import io
import json
import logging
import uuid
from pathlib import Path
from typing import Any

from fastapi import APIRouter, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from config import UPLOAD_DIR
from db import get_db, get_next_bd_code
from parsers.excel_parser import parse_excel, get_company_list, FIELD_LABELS, COLUMN_MAP

log = logging.getLogger(__name__)

router = APIRouter()


@router.get("/template")
async def download_template():
    """Generate and download an Excel template with all 26 column headers."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "标的信息模板"

    # Get all 26 Chinese column names from COLUMN_MAP
    headers = list(COLUMN_MAP.keys())

    # Write headers (bold)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Add example row
    example_data = {
        "标的编码": "BD001",
        "标的主体": "示例科技有限公司",
        "标的项目": "AI智能平台项目",
        "营业收入（元）": "50000000",
        "估值（元）": "200000000",
        "净利润（元）": "8000000",
        "行业": "人工智能",
        "上市编号": "",
        "估值日期": "2026-01-01",
        "官网地址": "https://example.com",
        "上市情况": "未上市",
        "标的描述": "专注于AI技术研发的高新技术企业",
        "标的主体公司简介": "成立于2020年，主营AI算法研发与应用",
        "省": "北京市",
        "市": "北京市",
        "区": "海淀区",
        "营业收入": "5000万",
        "净利润": "800万",
        "行业标签": "人工智能,大数据",
        "推介情况": "已推介",
        "是否已交易": "否",
        "负责人主属部门": "投资部",
        "归属部门": "投资部",
        "标的介绍材料附件": "",
        "标的公司年度报告摘要附件": "",
        "备注": "重点关注项目",
    }

    for col_idx, header in enumerate(headers, start=1):
        value = example_data.get(header, "")
        ws.cell(row=2, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(headers[col_idx - 1])
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    # Freeze first row
    ws.freeze_panes = "A2"

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=标的信息模板.xlsx"
        }
    )


# ── Session persistence helpers (SQLite) ─────────────────────

def _save_session(session_id: str, data: dict) -> None:
    """Persist session data to SQLite (parsed_texts stored as file refs)."""
    # Separate large parsed_texts to disk
    parsed_texts = data.get("parsed_texts", {})
    if parsed_texts:
        texts_dir = UPLOAD_DIR / session_id / "_parsed"
        texts_dir.mkdir(parents=True, exist_ok=True)
        refs: dict[str, list[dict]] = {}
        for bd_code, items in parsed_texts.items():
            refs[bd_code] = []
            for i, (filename, text) in enumerate(items):
                text_file = texts_dir / f"{bd_code}_{i}.txt"
                text_file.write_text(text, encoding="utf-8")
                refs[bd_code].append({"filename": filename, "path": str(text_file)})
        data_to_store = {**data, "parsed_texts_refs": refs}
    else:
        data_to_store = {**data}
    # Remove raw parsed_texts from JSON (too large)
    data_to_store.pop("parsed_texts", None)
    conn = get_db()
    try:
        conn.execute(
            "INSERT OR REPLACE INTO upload_sessions (session_id, data) VALUES (?, ?)",
            (session_id, json.dumps(data_to_store, ensure_ascii=False)),
        )
        conn.commit()
    finally:
        conn.close()


def _load_session(session_id: str) -> dict | None:
    """Load session data from SQLite, restoring parsed_texts from disk."""
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT data FROM upload_sessions WHERE session_id = ?",
            (session_id,),
        ).fetchone()
    finally:
        conn.close()
    if not row:
        return None
    data = json.loads(row["data"])
    # Restore parsed_texts from file refs
    refs = data.pop("parsed_texts_refs", {})
    parsed_texts: dict[str, list[tuple[str, str]]] = {}
    for bd_code, items in refs.items():
        parsed_texts[bd_code] = []
        for item in items:
            text_path = Path(item["path"])
            if text_path.exists():
                text = text_path.read_text(encoding="utf-8")
                parsed_texts[bd_code].append((item["filename"], text))
    data["parsed_texts"] = parsed_texts
    return data


# In-memory cache (backed by SQLite for persistence)
_sessions_cache: dict[str, dict] = {}


def get_session(session_id: str) -> dict:
    """Get session from cache or load from DB."""
    if session_id in _sessions_cache:
        return _sessions_cache[session_id]
    data = _load_session(session_id)
    if data is None:
        raise HTTPException(404, f"Session {session_id} not found")
    _sessions_cache[session_id] = data
    return data


def _persist_session(session_id: str, data: dict) -> None:
    """Save to both cache and DB."""
    _sessions_cache[session_id] = data
    _save_session(session_id, data)


class ManualInputRequest(BaseModel):
    """Manual input fields - company_name, project_name required; bd_code optional (auto-generated if empty)."""
    bd_code: str | None = None
    company_name: str
    project_name: str
    revenue_yuan: str | None = None
    valuation_yuan: str | None = None
    net_profit_yuan: str | None = None
    industry: str | None = None
    stock_code: str | None = None
    valuation_date: str | None = None
    website: str | None = None
    is_listed: str | None = None
    description: str | None = None
    company_intro: str | None = None
    province: str | None = None
    city: str | None = None
    district: str | None = None
    revenue: str | None = None
    net_profit: str | None = None
    industry_tags: str | None = None
    referral_status: str | None = None
    is_traded: str | None = None
    dept_primary: str | None = None
    dept_owner: str | None = None
    remarks: str | None = None


@router.post("/excel")
async def upload_excel(file: UploadFile = File(...)):
    """Upload an Excel file; returns parsed company list + session id."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Please upload an .xlsx or .xls file")

    session_id = uuid.uuid4().hex[:12]
    session_dir = UPLOAD_DIR / session_id
    session_dir.mkdir(parents=True, exist_ok=True)

    file_path = session_dir / file.filename
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    try:
        companies = get_company_list(str(file_path))
        all_rows = parse_excel(str(file_path))

        # Auto-generate bd_code for rows with empty bd_code
        for company in companies:
            if not company.get("bd_code") or not company["bd_code"].strip():
                company["bd_code"] = get_next_bd_code()
        for row in all_rows:
            if not row.get("bd_code") or not str(row["bd_code"]).strip():
                row["bd_code"] = get_next_bd_code()

    except Exception as e:
        raise HTTPException(422, f"Failed to parse Excel: {e}")

    session_data = {
        "excel_path": str(file_path),
        "companies": companies,
        "all_rows": all_rows,
        "attachments": {},
        "parsed_texts": {},
    }
    _persist_session(session_id, session_data)

    return {
        "session_id": session_id,
        "company_count": len(companies),
        "companies": companies,
    }


@router.post("/manual")
async def manual_input(req: ManualInputRequest):
    """Create a session from manually entered data (no Excel needed)."""
    session_id = uuid.uuid4().hex[:12]
    session_dir = UPLOAD_DIR / session_id
    session_dir.mkdir(parents=True, exist_ok=True)

    # Auto-generate bd_code if not provided
    bd_code = req.bd_code.strip() if req.bd_code else ""
    if not bd_code:
        bd_code = get_next_bd_code()

    row = req.model_dump()
    row = {k: v for k, v in row.items() if v is not None}
    row["bd_code"] = bd_code

    session_data = {
        "excel_path": None,
        "companies": [
            {
                "bd_code": bd_code,
                "company_name": req.company_name,
                "project_name": req.project_name,
            }
        ],
        "all_rows": [row],
        "attachments": {},
        "parsed_texts": {},
    }
    _persist_session(session_id, session_data)

    return {
        "session_id": session_id,
        "bd_code": bd_code,
        "company_name": req.company_name,
        "project_name": req.project_name,
    }


@router.get("/fields")
async def get_field_definitions():
    """Return the field definitions for manual input form."""
    required = ["company_name", "project_name"]  # bd_code now optional
    fields = []
    for en_key, cn_label in FIELD_LABELS.items():
        if en_key in ("intro_attachment", "annual_report_attachment"):
            continue
        fields.append({
            "key": en_key,
            "label": cn_label,
            "required": en_key in required,
        })
    return {"fields": fields}


def _parse_from_bytes(filename: str, raw_bytes: bytes) -> str:
    """Parse attachment text from raw in-memory bytes (before any disk encryption)."""
    ext = Path(filename).suffix.lower()

    if ext == ".pdf":
        return _parse_pdf_from_bytes(filename, raw_bytes)
    elif ext in (".md", ".txt"):
        return raw_bytes.decode("utf-8", errors="replace")
    elif ext == ".docx":
        return _parse_docx_from_bytes(filename, raw_bytes)
    elif ext == ".pptx":
        return _parse_pptx_from_bytes(filename, raw_bytes)
    return ""


def _parse_pdf_from_bytes(filename: str, raw_bytes: bytes) -> str:
    """Parse PDF from in-memory bytes using PyMuPDF, pdfplumber, and OCR."""
    text = ""

    # 1. Try PyMuPDF
    try:
        import fitz
        doc = fitz.open(stream=raw_bytes, filetype="pdf")
        parts = [page.get_text() for page in doc]
        page_count = doc.page_count
        doc.close()
        text = "\n\n".join(parts)
        log.info("  PyMuPDF (bytes): %d chars from %d pages", len(text.strip()), page_count)
    except Exception as e:
        log.debug("  PyMuPDF (bytes) failed for %s: %s", filename, e)
        page_count = 0

    # 2. Fallback: pdfplumber
    if len(text.strip()) < 100:
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
                page_count = len(pdf.pages)
                parts = []
                for page in pdf.pages:
                    pt = page.extract_text()
                    if pt:
                        parts.append(pt)
                text = "\n\n".join(parts)
                log.info("  pdfplumber (bytes): %d chars from %d pages", len(text.strip()), page_count)
        except Exception as e:
            log.debug("  pdfplumber (bytes) failed for %s: %s", filename, e)

    # 3. OCR if text extraction insufficient
    text_len = len(text.strip())
    need_ocr = (
        (page_count > 0 and text_len / page_count < 80)
        or (page_count == 0 and text_len < 100)
    )

    if need_ocr:
        log.info("  Text insufficient (%d chars), attempting OCR for %s...", text_len, filename)
        ocr_text = _ocr_pdf_from_bytes(raw_bytes, page_count)
        if ocr_text and len(ocr_text.strip()) > text_len:
            log.info("  OCR yielded %d chars (vs %d from text extraction)", len(ocr_text.strip()), text_len)
            text = ocr_text

    return text


def _ocr_pdf_from_bytes(raw_bytes: bytes, known_page_count: int) -> str:
    """OCR a PDF from in-memory bytes."""
    try:
        import fitz
        from rapidocr_onnxruntime import RapidOCR

        ocr_engine = RapidOCR()
        doc = fitz.open(stream=raw_bytes, filetype="pdf")
        max_pages = min(doc.page_count, 50)
        dpi = 200 if max_pages <= 20 else 150

        parts: list[str] = []
        for i in range(max_pages):
            pix = doc[i].get_pixmap(dpi=dpi)
            result, _ = ocr_engine(pix.tobytes("png"))
            if result:
                page_text = "\n".join([item[1] for item in result])
                if page_text.strip():
                    parts.append(page_text)
        doc.close()
        total = sum(len(p) for p in parts)
        log.info("  OCR (PyMuPDF bytes): %d pages → %d chars", max_pages, total)
        if total > 100:
            return "\n\n".join(parts)
    except Exception as e:
        log.debug("  OCR via PyMuPDF bytes failed: %s", e)

    try:
        import pdfplumber
        from rapidocr_onnxruntime import RapidOCR

        ocr_engine = RapidOCR()
        parts = []
        with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            max_pages = min(len(pdf.pages), 50)
            dpi = 200 if max_pages <= 20 else 150
            for i, page in enumerate(pdf.pages[:max_pages]):
                try:
                    img = page.to_image(resolution=dpi)
                    buf = io.BytesIO()
                    img.original.save(buf, format="PNG")
                    result, _ = ocr_engine(buf.getvalue())
                    if result:
                        page_text = "\n".join([item[1] for item in result])
                        if page_text.strip():
                            parts.append(page_text)
                except Exception:
                    pass
        total = sum(len(p) for p in parts)
        log.info("  OCR (pdfplumber bytes): %d pages → %d chars", max_pages, total)
        return "\n\n".join(parts)
    except Exception as e:
        log.debug("  OCR via pdfplumber bytes failed: %s", e)

    return ""


def _parse_docx_from_bytes(filename: str, raw_bytes: bytes) -> str:
    """Parse DOCX from in-memory bytes."""
    try:
        from parsers.docx_parser import extract_docx_text
        return extract_docx_text(io.BytesIO(raw_bytes))
    except Exception as e:
        log.debug("  DOCX bytes parse failed for %s: %s", filename, e)
        return ""


def _parse_pptx_from_bytes(filename: str, raw_bytes: bytes) -> str:
    """Parse PPTX from in-memory bytes."""
    try:
        from parsers.pptx_parser import extract_pptx_text
        return extract_pptx_text(io.BytesIO(raw_bytes))
    except Exception as e:
        log.debug("  PPTX bytes parse failed for %s: %s", filename, e)
        return ""


@router.post("/attachments")
async def upload_attachments(
    session_id: str,
    bd_code: str,
    files: list[UploadFile] = File(...),
):
    """Upload PDF/MD attachments for a specific company."""
    session = get_session(session_id)
    session_dir = UPLOAD_DIR / session_id / bd_code
    session_dir.mkdir(parents=True, exist_ok=True)

    saved: list[str] = []
    parsed: list[dict] = []
    skipped: list[str] = []

    for f in files:
        ext = Path(f.filename).suffix.lower()
        if ext not in (".pdf", ".md", ".txt", ".docx", ".pptx"):
            skipped.append(f.filename)
            continue

        raw_bytes = await f.read()
        if not raw_bytes:
            skipped.append(f.filename)
            continue

        dest = session_dir / f.filename
        with open(dest, "wb") as out:
            out.write(raw_bytes)
        saved.append(str(dest))

        log.info("Parsing attachment from memory: %s (%d bytes)", f.filename, len(raw_bytes))
        try:
            text = _parse_from_bytes(f.filename, raw_bytes)
            text_len = len(text.strip()) if text else 0
            log.info("Parsed %s: %d chars", f.filename, text_len)
            parsed.append({"filename": f.filename, "text_length": text_len})

            if text and text.strip():
                session.setdefault("parsed_texts", {}).setdefault(bd_code, []).append(
                    (f.filename, text)
                )
            else:
                log.warning("Attachment parsed but empty: %s", f.filename)
        except Exception as e:
            log.exception("Failed to parse attachment from memory: %s", f.filename)
            parsed.append({"filename": f.filename, "text_length": 0, "error": str(e)})

    session["attachments"].setdefault(bd_code, []).extend(saved)
    # Persist updated session
    _persist_session(session_id, session)

    return {
        "bd_code": bd_code,
        "uploaded": len(saved),
        "files": saved,
        "parsed": parsed,
        "skipped": skipped,
    }


@router.get("/session/{session_id}")
async def get_session_info(session_id: str):
    """Return session metadata (companies, attachments)."""
    session = get_session(session_id)
    return {
        "session_id": session_id,
        "companies": session["companies"],
        "attachments": {k: len(v) for k, v in session["attachments"].items()},
    }


# Expose sessions for other routers
def get_sessions():
    return _sessions_cache
